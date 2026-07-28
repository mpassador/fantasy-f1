from collections import defaultdict
from openpyxl import load_workbook

FILE = "/Users/moa.passador/repos/fantasy-f1/app/Formula 1.xlsx"

# Standard F1 points
RACE_POINTS = {1:25, 2:18, 3:15, 4:12, 5:10, 6:8, 7:6, 8:4, 9:2, 10:1}
SPRINT_POINTS = {1:8, 2:7, 3:6, 4:5, 5:4, 6:3, 7:2, 8:1}

wb = load_workbook(FILE, data_only=True)
summary = wb["2026"]
results = wb["Resultados 2026"]

# ------------------------------------------------------------
# Players (rows 6-14 in the 2026 sheet)
# ------------------------------------------------------------
players = []
for row in range(6, 15):
    players.append({
        "name": summary[f"B{row}"].value,
        "driver": summary[f"C{row}"].value.lower(),
        "team": summary[f"D{row}"].value.lower(),
        "expected_total": summary[f"AH{row}"].value,
    })

# ------------------------------------------------------------
# Helper: normalize names
# ------------------------------------------------------------
def surname(full_name: str) -> str:
    return full_name.split()[-1].lower()

def normalize_team(team: str) -> str:
    team = team.lower()
    if "red bull" in team:
        return "red bull"
    if "haas" in team:
        return "haas"
    return team

# ------------------------------------------------------------
# Collect points from every race
# ------------------------------------------------------------
driver_points = defaultdict(lambda: defaultdict(int))
team_points = defaultdict(lambda: defaultdict(int))

# ---- First four races (wide layout) ----
# "Australia" uses a different layout: driver+team are combined into a
# single cell (e.g. "G. Russell\nMercedes\n·#63"), with no separate
# Team column. The other three races have separate Driver/Team columns.
wide_races = [
    ("Australia", 1, 2, None, 4),       # pos_col, driver_col (combined), team_col, pts_col
    ("China Sprint", 5, 7, 8, 10),
    ("China Corrida", 11, 13, 14, 16),
    ("Japao", 17, 19, 20, 22),
]

for race_name, pos_col, driver_col, team_col, pts_col in wide_races:
    for row in range(3, 13):
        pts = results.cell(row, pts_col).value or 0

        if team_col is None:
            # Combined driver/team cell (Australia)
            cell_value = results.cell(row, driver_col).value
            if cell_value:
                parts = [p.strip() for p in cell_value.split("\n") if p.strip()]
                drv = parts[0] if len(parts) > 0 else None
                team = parts[1] if len(parts) > 1 else None
            else:
                drv = team = None
        else:
            drv = results.cell(row, driver_col).value
            team = results.cell(row, team_col).value

        if drv:
            driver_points[race_name][surname(drv)] = int(pts)
        if team:
            team_points[race_name][normalize_team(team)] += int(pts)

# ---- Remaining races (vertical layout) ----
row = 14
while row <= results.max_row:
    race_name = results.cell(row, 1).value

    if race_name and race_name != "Pos.":
        is_sprint = "sprint" in race_name.lower()
        points_table = SPRINT_POINTS if is_sprint else RACE_POINTS

        data_row = row + 2
        while data_row <= results.max_row:
            pos = results.cell(data_row, 1).value

            if not isinstance(pos, (int, float)):
                break

            drv = results.cell(data_row, 3).value
            team = results.cell(data_row, 4).value
            pts = points_table.get(int(pos), 0)

            if drv:
                driver_points[race_name][surname(drv)] = pts
            if team:
                team_points[race_name][normalize_team(team)] += pts

            data_row += 1

        row = data_row
    else:
        row += 1

# ------------------------------------------------------------
# Calculate standings
# ------------------------------------------------------------
standings = []

all_races = set(driver_points.keys()) | set(team_points.keys())

for p in players:
    total = 0

    for race in all_races:
        total += driver_points[race].get(p["driver"], 0)
        total += team_points[race].get(p["team"], 0)

    standings.append({
        "name": p["name"],
        "driver": p["driver"],
        "team": p["team"],
        "python_total": total,
        "excel_total": p["expected_total"],
        "match": total == p["expected_total"],
    })

standings.sort(key=lambda x: x["python_total"], reverse=True)

# ------------------------------------------------------------
# Print results
# ------------------------------------------------------------
print(f"{'Player':12} {'Python':>6} {'Excel':>6} {'OK':>4}")
print("-" * 34)

for s in standings:
    ok = "✓" if s["match"] else "✗"
    print(f"{s['name']:12} {s['python_total']:6} {s['excel_total']:6} {ok:>4}")